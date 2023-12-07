from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options # Opçoes de inicialização do Chrome
from time import sleep
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def iniciar_driver():
    chrome_options = Options()
    argumentos = ['--leng=pt-BR', '--window-size=1000,700', '--incognito']

    for argumento in argumentos:
        chrome_options.add_argument(argumento)
        
    # Uso de configurações experimentais
    chrome_options.add_experimental_option('prefs', {
        # Desabilitar a confirmação de download
        'download.prompt_for_download': False,
        # Desabilitar notificações
        'profile.default_content_setting_values.notifications': 2,
        # Permitir multiplos downloads
        'profile.default_content_setting_values.automatic_downloads': 1,

    })

    # Inicializando o webdriver
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)
    
    return driver


'''
- Abrir site
- Clicar no primeiro "Clique aqui"
> Este fluxo deve se repetir pra cada estado do brasil
	- Clicar em um estado
	- Clicar em uma cidade
    - Verificar cada cartorio da cidade
	- Pegar nome do cartorio, responsável, telefone, e-mail e atribuição de cada cartório
- Salvar em uma planilha Excel
'''

def inserir_dados_planilha(denominacao, responsavel, atribuicoes, endereco, telefone_e_email, estado_atual):

    # Cria ou recarrega o arquivo
    try:
        workbook = load_workbook('cartorios.xlsx')
    except FileNotFoundError:
        # Se o arquivo nao existir, cria um novo
        workbook = Workbook()
    
    # Seleciona ou cria uma planilha por estado
    if estado_atual not in workbook.sheetnames:
        sheet = workbook.create_sheet(title=estado_atual)
    else:
        sheet = workbook[estado_atual]
    
    # Criando células do header se a planilha estiver vazia
    if sheet.max_row == 1:
        # Criando cedulas do header
        sheet['A1'] = 'Denominção'
        sheet['B1'] = 'Responsável'
        sheet['C1'] = 'Atribuições'
        sheet['D1'] = 'Endereço'
        sheet['E1'] = 'Telefone e E-mail'
        
        # inserindo estilo ao header
        for coluna in ['A', 'B', 'C', 'D', 'E']:
            header = sheet[f'{coluna}1']
            header.font = Font(color="FFFFFF", bold=True)
            header.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        
    sheet.append([denominacao, responsavel, atribuicoes, endereco, telefone_e_email])
    
    workbook.save('cartorios.xlsx')


def clicar_elemento(driver, by, value):
    elemento = driver.find_element(by, value)
    driver.execute_script('arguments[0].click()', elemento)
    

def obter_dados_cartorio(driver, estado_atual_main):
    try:
        # Colocando para mostrar o maximo de registros - NAO CLICOU, TESTAR MAIS
        # clicar_elemento(driver, By.XPATH, "(//b[normalize-space()='AC'])[1]")
        # clicar_elemento(driver, By.XPATH, '//*[@id="display_length"]/label/select/option[4]')
        
        sleep(5)
        
        # Verificando a quantidade de cartorios
        quantidade_de_registros = driver.find_elements(By.XPATH, "//tr[contains(@class, 'processo')]")
        
        # Verificando a cidade atual
        cidade_atual = driver.find_element(By.XPATH, "/html/body/div[2]/div[5]/fieldset/legend/b[1]").text
        print(f'São {len(quantidade_de_registros)} registros para a cidade de {cidade_atual}')
        
        estado_atual = driver.find_element(By.XPATH, f'(//b[normalize-space()=\"{estado_atual_main}\"])[1]').text # (//b[normalize-space()='AC'])[1]
        print(f'Estado atual: {estado_atual}')
        
        print('\n\n')
        
        sleep(2)
        
    
        for registro in quantidade_de_registros:
            # Pegando os dados dos cartórios
            try:
                denominacao = driver.find_element(By.XPATH, ".//td[2]/table/tbody/tr[1]/td[2]").text
                print(f'Denominação: {denominacao}')
            except Exception as erro:
                print(f'Erro ao encontrar elemento {erro}')
                denominacao = 'Sem Denominação'
            
            try:
                responsavel = driver.find_element(By.XPATH, ".//td[2]/table/tbody/tr[2]/td[2]").text
                print(f'Responsável: {responsavel}')
            except Exception as erro:
                print(f'Erro ao encontrar elemento {erro}')
                responsavel = 'Sem Responsável'
            
            try:
                atribuicoes = driver.find_element(By.XPATH, ".//td[2]/table/tbody/tr[3]/td[2]").text
                print(f'Atribuições: {atribuicoes}')
            except Exception as erro:
                print(f'Erro ao encontrar elemento {erro}')
                atribuicoes = 'Sem Atribuições'
            
            try:
                endereco = driver.find_element(By.XPATH, ".//td[2]/table/tbody/tr[4]/td[2]").text
                print(f'Endereço: {endereco}')
            except Exception as erro:
                print(f'Erro ao encontrar elemento {erro}')
                endereco = 'Sem Endereço'
            
            try:
                telefone_e_email = driver.find_element(By.XPATH, ".//td[2]/table/tbody/tr[5]/td[2]").text
                print(f'Telefone e E-mail: {telefone_e_email}')
            except Exception as erro:
                print(f'Erro ao encontrar elemento {erro}')
                telefone_e_email = 'Sem Telefone e E-mail'
            
            print('\n\n')
            
            # Chamando a funcao para inserir os dados na planilha
            inserir_dados_planilha(denominacao, responsavel, atribuicoes, endereco, telefone_e_email, estado_atual)
        
        # Verificando se o campo "Seguinte" esta habilitado para mostrar mais cartorios ou nao
        try:
            campo_seguinte = driver.find_element(By.XPATH, '//a[@class="next fg-button ui-button ui-state-default"]')
            
            # Verifica se o campo esta clicavel
            elemento = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(campo_seguinte))
            sleep(2)
            clicar_elemento(driver, By.XPATH, '//a[@class="next fg-button ui-button ui-state-default"]')
            sleep(2)
            # Se estiver clicavel, chama a funcao novamente para obter os dados
            obter_dados_cartorio(driver)
        except NoSuchElementException:
            print('Botão "Seguinte" não encontrado. Avançando para a próxima cidade.')
        
        
    except Exception as erro:
        print(f'Erro ao obter dados do cartório: {erro}')


def processar_estado(driver, estado):
    # Entrando no estado
    clicar_elemento(driver, By.XPATH, f'//area[contains(@onclick, "pesquisaServentiasExtra(\'{estado}\')")]')
    print(f'Entrando no estado de {estado}')
    estado_atual_main = estado

    sleep(1)

    # Verificando quantas cidades tem cada estado
    quantidade_de_cidades = driver.find_elements(By.XPATH, '//option[@value]')
    print(f'O estado de {estado} possui {len(quantidade_de_cidades) - 1} cidades')

    # Entrando em cada cidade de um estado
    for cidade in range(1, len(quantidade_de_cidades)):
        sleep(1)
        btn_cidade = driver.find_elements(By.XPATH, '//option[@value]')
        sleep(3)
        # Clicando na cidade
        btn_cidade[cidade].click()
        sleep(1)

        # Clicando no botao pesquisar cidades para ver os cartorios
        clicar_elemento(driver, By.XPATH, '//*[@id="div_cidade"]/div/table/tbody/tr[2]/td/button[1]')

        sleep(1)

        # Chamando a funcao para obter os dados de cada cartorio
        obter_dados_cartorio(driver, estado_atual_main)
        
        sleep(1)
        # Clicando em Extrajudicial para abrir o dropdown
        clicar_elemento(driver, By.XPATH, "(//a[normalize-space()='Extrajudicial'])[1]")
        sleep(1)
        
        # Clicando em Serventias Extrajudiciais para voltar para os estados
        clicar_elemento(driver, By.XPATH, "(//a[normalize-space()='Serventias Extrajudiciais'])[1]")
        sleep(1)
        
        # Clicando novamente no estado
        clicar_elemento(driver, By.XPATH, f'//area[contains(@onclick, "pesquisaServentiasExtra(\'{estado}\')")]')


    sleep(1)
    clicar_elemento(driver, By.XPATH, "(//a[normalize-space()='Extrajudicial'])[1]")
    sleep(1)


def main():
    # Chamando a funcao para iniciar o driver
    driver = iniciar_driver()
    
    # Navegar até um site
    driver.get('https://www.cnj.jus.br/corregedoria/justica_aberta/?')

    # Clicar no primeiro "clique aqui"
    clicar_elemento(driver, By.XPATH, '//strong[text()="Clique aqui"]')
    sleep(3)

    # Estados brasileiros
    siglas_estados_brasileiros = [
        'AC', 'AL', 'AP', 'AM', 'BA', 'CE', 'DF', 'ES', 'GO', 'MA',
        'MT', 'MS', 'MG', 'PA', 'PB', 'PR', 'PE', 'PI', 'RJ', 'RN',
        'RS', 'RO', 'RR', 'SC', 'SP', 'SE', 'TO'
    ]

    for estado in siglas_estados_brasileiros:
        processar_estado(driver, estado)


    input('Pressione uma tecla para fechar: ')
    driver.quit()


if __name__ == "__main__":
    main()
